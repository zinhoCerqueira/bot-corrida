#!/usr/bin/env python3
"""
GERADOR DE BACKUP/SETUP DO PLANO (EXCEL)
Finalidade: Este script é utilizado APENAS para:
1. Setup Inicial: Criar a estrutura base da planilha.
2. Backup Offline: Gerar um arquivo .xlsx local com a formatação original.

IMPORTANTE: Este script NÃO sincroniza dados realizados. Para sincronizar o log 
de treinos do Markdown para o Google Sheets, use o 'scripts/sync_to_sheets.py'.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re

OUTPUT = '/home/zzz/Documentos/Plano_Treino_2026.xlsx'

def pace_to_decimal(pace_str):
    """Converte pace string (ex: '6:30', '6:30–6:45') em minutos decimais."""
    if not pace_str:
        return None
    matches = re.findall(r'(\d+):(\d+)', pace_str)
    if not matches:
        return None
    decimals = [int(m) + int(s)/60.0 for m, s in matches]
    return sum(decimals) / len(decimals)

# ─── Dados de Treinamento ──────────────────────────────────────────────────────
# (data, dia, semana, fase, tipo, dist_km, pace_alvo, zona, detalhes, semana_rec, pace_real, percepcao, comentarios)
sessions = [
    ('2026-04-30','Qui', 0,'Aquecimento', 'CL',   6.0,'6:35–6:50','Z1',    'Recuperação pós-tiros',False, ' ', None, ' '),
    # ── Fase 1: Base Aeróbia ──
    ('2026-05-03','Dom', 1,'Base Aeróbia','LR',  10.0,'6:30–6:45','Z1-Z2', 'Corrida longa fácil',False, ' ', None, ' '),
    ('2026-05-05','Ter', 1,'Base Aeróbia','FK',   7.0,'Z2–Z4',    'Z2-Z4', '2km Z1 · 5×(2min Z4/2min Z1) · 1km Z1',False, ' ', None, ' '),
    ('2026-05-07','Qui', 1,'Base Aeróbia','CL',   6.0,'6:35–6:50','Z1',    'Corrida leve',False, ' ', None, ' '),
    ('2026-05-10','Dom', 2,'Base Aeróbia','LR',  11.0,'6:30–6:40','Z1-Z2', 'Corrida longa',False, ' ', None, ' '),
    ('2026-05-12','Ter', 2,'Base Aeróbia','PROG', 7.0,'Z2→Z3',    'Z2-Z3', '3km Z2 → 2km Z2/Z3 → 2km Z3',False, ' ', None, ' '),
    ('2026-05-14','Qui', 2,'Base Aeróbia','CL',   6.0,'6:35–6:50','Z1',    'Corrida leve',False, ' ', None, ' '),
    ('2026-05-17','Dom', 3,'Base Aeróbia','LR',  12.0,'6:30–6:40','Z1-Z2', 'Corrida longa',False, ' ', None, ' '),
    ('2026-05-19','Ter', 3,'Base Aeróbia','FK',   8.0,'Z2–Z4',    'Z2-Z4', '2km Z1 · 6×(2min Z4/90s Z1) · 2km Z1',False, ' ', None, ' '),
    ('2026-05-21','Qui', 3,'Base Aeróbia','CM',   7.0,'6:05–6:20','Z2',    'Corrida moderada',False, ' ', None, ' '),
    ('2026-05-24','Dom', 4,'Base Aeróbia','LR',   9.0,'6:40–6:50','Z1',    'LR — semana leve ⚡',True, ' ', None, ' '),
    ('2026-05-26','Ter', 4,'Base Aeróbia','CL',   5.0,'6:40–6:50','Z1',    'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-05-28','Qui', 4,'Base Aeróbia','CL',   5.0,'6:40–6:50','Z1',    'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-05-31','Dom', 5,'Base Aeróbia','LR',  12.0,'6:30–6:40','Z1-Z2', 'Corrida longa',False, ' ', None, ' '),
    ('2026-06-02','Ter', 5,'Base Aeróbia','TM',   8.0,'5:00–5:08','Z4',    '2km aquec · 4×1km Z4 c/90s rec · 2km cool',False, ' ', None, ' '),
    ('2026-06-04','Qui', 5,'Base Aeróbia','CL',   7.0,'6:40',     'Z1',    'Corrida leve',False, ' ', None, ' '),
    ('2026-06-07','Dom', 6,'Base Aeróbia','LR',  13.0,'6:30–6:40','Z1-Z2', 'Corrida longa',False, ' ', None, ' '),
    ('2026-06-09','Ter', 6,'Base Aeróbia','PROG', 8.0,'Z2→Z3',    'Z2-Z3', '3km Z2 → 2km Z3 → 2km Z3 forte → 1km Z2',False, ' ', None, ' '),
    ('2026-06-11','Qui', 6,'Base Aeróbia','CM',   7.0,'6:05–6:15','Z2',    'Corrida moderada',False, ' ', None, ' '),
    ('2026-06-14','Dom', 7,'Base Aeróbia','LR',  14.0,'6:30–6:40','Z1-Z2', 'Corrida longa',False, ' ', None, ' '),
    ('2026-06-16','Ter', 7,'Base Aeróbia','TM',   9.0,'5:00–5:05','Z4',    '2km aquec · 5×1km Z4 c/90s rec · 2km cool',False, ' ', None, ' '),
    ('2026-06-18','Qui', 7,'Base Aeróbia','CL',   7.0,'6:40',     'Z1',    'Corrida leve',False, ' ', None, ' '),
    ('2026-06-21','Dom', 8,'Base Aeróbia','LR',  10.0,'6:40–6:50','Z1',    'LR — semana leve ⚡',True, ' ', None, ' '),
    ('2026-06-23','Ter', 8,'Base Aeróbia','CL',   5.0,'6:40–6:50','Z1',    'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-06-25','Qui', 8,'Base Aeróbia','CL',   5.0,'6:40–6:50','Z1',    'Leve — recuperação ⚡',True, ' ', None, ' '),
    # ── Fase 2: Desenvolvimento ──
    ('2026-06-28','Dom', 9,'Desenvolvimento','LR', 14.0,'6:28–6:38','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-06-30','Ter', 9,'Desenvolvimento','TR',  8.0,'5:27–5:32','Z3',   '2km aquec · 25min Z3 · 2km cool',False, ' ', None, ' '),
    ('2026-07-02','Qui', 9,'Desenvolvimento','CM',  7.0,'6:08–6:18','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-07-05','Dom',10,'Desenvolvimento','LR', 15.0,'6:28',     'Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-07-07','Ter',10,'Desenvolvimento','TM',  9.0,'5:00–5:05','Z4',   '2km aquec · 5×1km Z4 c/90s rec · 2km cool',False, ' ', None, ' '),
    ('2026-07-09','Qui',10,'Desenvolvimento','CM',  8.0,'6:05–6:15','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-07-12','Dom',11,'Desenvolvimento','LR', 16.0,'6:25–6:35','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-07-14','Ter',11,'Desenvolvimento','TR',  9.0,'5:27–5:32','Z3',   '2km aquec · 30min Z3 · 2km cool',False, ' ', None, ' '),
    ('2026-07-16','Qui',11,'Desenvolvimento','CL+S',7.0,'6:40 + strides','Z1+Z5','6km Z1 · 4×20s aceleração',False, ' ', None, ' '),
    ('2026-07-19','Dom',12,'Desenvolvimento','LR', 12.0,'6:40–6:50','Z1',   'LR — semana leve ⚡',True, ' ', None, ' '),
    ('2026-07-21','Ter',12,'Desenvolvimento','CL',  6.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-07-23','Qui',12,'Desenvolvimento','CL',  6.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-07-26','Dom',13,'Desenvolvimento','LR', 16.0,'6:25–6:35','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-07-28','Ter',13,'Desenvolvimento','TM', 10.0,'5:00–5:05','Z4',   '2km aquec · 6×1km Z4 c/90s rec · 2km cool',False, ' ', None, ' '),
    ('2026-07-30','Qui',13,'Desenvolvimento','CM',  8.0,'6:05–6:15','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-08-02','Dom',14,'Desenvolvimento','LR', 17.0,'6:25–6:35','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-08-04','Ter',14,'Desenvolvimento','TR', 10.0,'5:25–5:30','Z3',   '2km aquec · 35min Z3 · 2km cool',False, ' ', None, ' '),
    ('2026-08-06','Qui',14,'Desenvolvimento','CM',  8.0,'6:05–6:15','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-08-09','Dom',15,'Desenvolvimento','LR', 17.0,'6:25–6:35','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-08-11','Ter',15,'Desenvolvimento','TM', 10.0,'5:02–5:08','Z4',   '2km aquec · 4×1,5km Z4 c/2min rec · 2km cool',False, ' ', None, ' '),
    ('2026-08-13','Qui',15,'Desenvolvimento','CL+S',8.0,'6:40 + strides','Z1+Z5','7km Z1 · 4×20s aceleração',False, ' ', None, ' '),
    ('2026-08-16','Dom',16,'Desenvolvimento','LR', 13.0,'6:40–6:50','Z1',   'LR — semana leve ⚡',True, ' ', None, ' '),
    ('2026-08-18','Ter',16,'Desenvolvimento','CL',  6.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-08-20','Qui',16,'Desenvolvimento','CL',  6.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-08-23','Dom',17,'Desenvolvimento','LR', 18.0,'6:22–6:32','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-08-25','Ter',17,'Desenvolvimento','TR', 10.0,'5:25–5:30','Z3',   '2km aquec · 35min Z3 · 2km cool',False, ' ', None, ' '),
    ('2026-08-27','Qui',17,'Desenvolvimento','CM',  9.0,'6:05–6:15','Z2',   'Corrida moderada',False, ' ', None, ' '),
    # ── Fase 3: Específico ──
    ('2026-08-30','Dom',18,'Específico','RETESTE', 5.0,'5:00–5:10','Z5',   'Teste 5km (Recalibragem) + 3km (aq/des)',False, ' ', None, ' '),
    ('2026-09-01','Ter',18,'Específico','TM', 10.0,'4:58–5:05','Z4',   '2km aquec · 6×1km Z4 c/90s rec · 2km cool',False, ' ', None, ' '),
    ('2026-09-03','Qui',18,'Específico','CM',  9.0,'6:05–6:15','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-09-06','Dom',19,'Específico','LR', 16.0,'6:28',     'Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-09-08','Ter',19,'Específico','TR', 11.0,'5:25–5:30','Z3',   '2km aquec · 40min Z3 · 2km cool',False, ' ', None, ' '),
    ('2026-09-10','Qui',19,'Específico','CM',  9.0,'6:05–6:15','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-09-13','Dom',20,'Específico','LR', 19.0,'6:20–6:30','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-09-15','Ter',20,'Específico','TM', 10.0,'5:05–5:10','Z4',   '2km aquec · 3×2km Z4 c/2min rec · 2km cool',False, ' ', None, ' '),
    ('2026-09-17','Qui',20,'Específico','CL+S',8.0,'6:40 + strides','Z1+Z5','7km Z1 · 5×20s aceleração',False, ' ', None, ' '),
    ('2026-09-20','Dom',21,'Específico','LR', 14.0,'6:40–6:50','Z1',   'LR — semana leve ⚡',True, ' ', None, ' '),
    ('2026-09-22','Ter',21,'Específico','CL',  7.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-09-24','Qui',21,'Específico','CL',  7.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-09-27','Dom',22,'Específico','LR', 19.0,'6:20–6:30','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-09-29','Ter',22,'Específico','RP', 12.0,'5:35–5:40','Z3',   '2km aquec · 5×2km RP c/90s rec · 2km cool',False, ' ', None, ' '),
    ('2026-10-01','Qui',22,'Específico','CM',  9.0,'6:05–6:15','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-10-04','Dom',23,'Específico','LR', 18.0,'6:22–6:32','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-10-06','Ter',23,'Específico','TR', 11.0,'5:22–5:27','Z3',   '2km aquec · 40min Z3 · 2km cool',False, ' ', None, ' '),
    ('2026-10-08','Qui',23,'Específico','CM',  9.0,'6:05–6:15','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-10-11','Dom',24,'Específico','LR', 20.0,'6:20–6:30','Z1-Z2','Corrida longa — marco 20km!',False, ' ', None, ' '),
    ('2026-10-13','Ter',24,'Específico','TM', 10.0,'5:02–5:07','Z4',   '2km aquec · 3×2km Z4 c/2min rec · 2km cool',False, ' ', None, ' '),
    ('2026-10-15','Qui',24,'Específico','CL+S',9.0,'6:40 + strides','Z1+Z5','8km Z1 · 5×20s aceleração',False, ' ', None, ' '),
    ('2026-10-18','Dom',25,'Específico','LR', 14.0,'6:40–6:50','Z1',   'LR — semana leve ⚡',True, ' ', None, ' '),
    ('2026-10-20','Ter',25,'Específico','CL',  7.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-10-22','Qui',25,'Específico','CL',  7.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    # ── Fase 4: Pico ──
    ('2026-10-25','Dom',26,'Pico','LR', 20.0,'6:20–6:30','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-10-27','Ter',26,'Pico','RP', 13.0,'5:35',     'Z3',   '2km aquec · 6×2km RP c/90s rec · 2km cool',False, ' ', None, ' '),
    ('2026-10-29','Qui',26,'Pico','CM', 10.0,'6:05–6:12','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-11-01','Dom',27,'Pico','LR', 19.0,'6:20–6:30','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-11-03','Ter',27,'Pico','TR', 12.0,'5:20–5:25','Z3',   '2km aquec · 45min Z3 · 2km cool',False, ' ', None, ' '),
    ('2026-11-05','Qui',27,'Pico','CL+S',9.0,'6:40 + strides','Z1+Z5','8km Z1 · 5×20s aceleração',False, ' ', None, ' '),
    ('2026-11-08','Dom',28,'Pico','LR', 20.0,'6:18–6:28','Z1-Z2','Corrida longa — última pesada!',False, ' ', None, ' '),
    ('2026-11-10','Ter',28,'Pico','RP', 12.0,'5:35',     'Z3',   '2km aquec · 3×3km RP c/2min rec · 2km cool',False, ' ', None, ' '),
    ('2026-11-12','Qui',28,'Pico','CM', 10.0,'6:03–6:12','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-11-15','Dom',29,'Pico','LR', 15.0,'6:40–6:50','Z1',   'LR — semana leve ⚡',True, ' ', None, ' '),
    ('2026-11-17','Ter',29,'Pico','CL',  8.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    ('2026-11-19','Qui',29,'Pico','CL',  7.0,'6:40–6:50','Z1',   'Leve — recuperação ⚡',True, ' ', None, ' '),
    # ── Fase 5: Taper ──
    ('2026-11-22','Dom',30,'Taper','RETESTE', 5.0,'4:55–5:05','Z5',   'Teste 5km (Recalibragem) + 3km (aq/des)',False, ' ', None, ' '),
    ('2026-11-24','Ter',30,'Taper','TM',  8.0,'4:58–5:05','Z4',   '2km aquec · 4×1km Z4 c/90s rec · 2km cool',False, ' ', None, ' '),
    ('2026-11-26','Qui',30,'Taper','CM',  8.0,'6:08–6:18','Z2',   'Corrida moderada',False, ' ', None, ' '),
    ('2026-11-29','Dom',31,'Taper','LR', 15.0,'6:30–6:40','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-12-01','Ter',31,'Taper','RP+TM',10.0,'5:40+5:10','Z3+Z4','2km aquec · 3×2km RP · 2×1km Z4 · 2km cool',False, ' ', None, ' '),
    ('2026-12-03','Qui',31,'Taper','CL',  7.0,'6:40–6:50','Z1',   'Corrida leve',False, ' ', None, ' '),
    ('2026-12-06','Dom',32,'Taper','LR', 13.0,'6:32–6:42','Z1-Z2','Corrida longa',False, ' ', None, ' '),
    ('2026-12-08','Ter',32,'Taper','TR',  8.0,'5:25–5:30','Z3',   '2km aquec · 25min Z3 · 2km cool',False, ' ', None, ' '),
    ('2026-12-10','Qui',32,'Taper','CL',  7.0,'6:40–6:50','Z1',   'Corrida leve',False, ' ', None, ' '),
    ('2026-12-13','Dom',33,'Taper','CL', 10.0,'6:40–6:50','Z1',   'Corrida leve',False, ' ', None, ' '),
    ('2026-12-15','Ter',33,'Taper','CL+S',6.0,'6:40 + strides','Z1+Z5','5km Z1 · 4×20s aceleração suave',False, ' ', None, ' '),
    ('2026-12-17','Qui',33,'Taper','CL',  5.0,'6:40–6:50','Z1',   'Corrida leve',False, ' ', None, ' '),
    ('2026-12-20','Dom',34,'Taper','CL',  8.0,'6:40',     'Z1',   'Pernas frescas — ritmo livre',False, ' ', None, ' '),
    ('2026-12-22','Ter',34,'Taper','CL+S',5.0,'6:40 + strides','Z1+Z5','4km Z1 · 4×20s aceleração leve',False, ' ', None, ' '),
    ('2026-12-24','Qui',34,'Taper','CL',  4.0,'6:40',     'Z1',   'Opcional — descanse se sentir pesado',False, ' ', None, ' '),
    # ── Prova ──
    ('2026-12-27','Dom',35,'Prova',    'PROVA',21.1,'5:40','Z3','MEIA MARATONA — Sub-2h00',False, ' ', None, ' '),
    # ── Recuperação ──
    ('2026-12-29','Ter',35,'Recuperação','REC',5.0,'7:00+','Z1','Trote leve — recuperação ativa',False, ' ', None, ' '),
    ('2026-12-31','Qui',35,'Recuperação','REC',5.0,'7:00+','Z1','Trote leve — celebre o ano!',False, ' ', None, ' '),
]

# ─── Color maps ───────────────────────────────────────────────────────────────
FASE_BG = {
    'Aquecimento':  'FFF9C4',
    'Base Aeróbia': 'DDEEFF',
    'Desenvolvimento': 'EDE7F6',
    'Específico':   'FFF3E0',
    'Pico':         'FCE4EC',
    'Taper':        'E8F5E9',
    'Prova':        'FF8A80',
    'Recuperação':  'F5F5F5',
}
TIPO_BG = {
    'LR':    'BBDEFB',
    'CL':    'C8E6C9',
    'CM':    'DCEDC8',
    'FK':    'E1BEE7',
    'PROG':  'B2EBF2',
    'TM':    'FFCCBC',
    'TC':    'FFAB91',
    'TR':    'FFE0B2',
    'RP':    'FFCDD2',
    'CL+S':  'B2DFDB',
    'RP+TM': '4A148C',
    'PROVA': 'FF5252',
    'RETESTE': '2196F3',
    'REC':   'EEEEEE',
}
TIPO_FG = {
    'PROVA': 'FFFFFF',
    'RP+TM': 'FFFFFF',
    'RETESTE': 'FFFFFF',
}

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def hdr_font(color='FFFFFF', bold=True, size=11):
    return Font(bold=bold, color=color, size=size)

def cell_font(bold=False, size=10, color='000000'):
    return Font(bold=bold, size=size, color=color)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 1 — PLANO COMPLETO
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Plano Completo'
ws1.freeze_panes = 'A3'

# Title row
ws1.merge_cells('A1:M1')
t = ws1['A1']
t.value = 'PLANO DE TREINAMENTO 2026 — MEIA MARATONA SUB-2H'
t.font = Font(bold=True, size=14, color='FFFFFF')
t.fill = fill('1A3A6B')
t.alignment = center()
ws1.row_dimensions[1].height = 28

# Header row
headers = ['Data', 'Dia', 'Semana', 'Fase', 'Tipo', 'Distância (km)', 'Pace Alvo', 'Pace Alvo Médio', 'Zona', 'Detalhes', 'Pace Real', 'Percepção (0-10)', 'Comentários']
col_widths = [13, 6, 9, 17, 8, 16, 14, 16, 10, 52, 12, 18, 40]
for c, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(2, c, h)
    cell.fill = fill('2B5BA8')
    cell.font = hdr_font()
    cell.alignment = center()
    cell.border = thin_border()
    ws1.column_dimensions[get_column_letter(c)].width = w
ws1.row_dimensions[2].height = 22

# Data rows
for r, s in enumerate(sessions, 3):
    date_str, dia, sem, fase, tipo, dist, pace, zona, det, rec, pace_real, perc, com = s
    pace_medio = pace_to_decimal(pace)
    row_vals = [date_str, dia, f'S{sem:02d}' if sem > 0 else '—', fase, tipo, dist, pace, pace_medio, zona, det, pace_real, perc, com]

    fase_color = FASE_BG.get(fase, 'FFFFFF')
    tipo_color = TIPO_BG.get(tipo, 'FFFFFF')
    row_bg = 'FFF8E1' if rec else fase_color  # recovery weeks get light yellow

    for c, val in enumerate(row_vals, 1):
        cell = ws1.cell(r, c, val)
        cell.border = thin_border()
        cell.font = cell_font()

        if c == 5:  # Tipo column — use workout color
            cell.fill = fill(tipo_color)
            cell.font = Font(bold=True, size=10, color=TIPO_FG.get(tipo, '000000'))
            cell.alignment = center()
        elif c == 6:  # Distance — right-aligned number
            cell.number_format = '0.0'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill(row_bg)
        elif c == 8:  # Pace Alvo Médio — numeric 2 decimals
            cell.number_format = '0.00'
            cell.alignment = center()
            cell.fill = fill(row_bg)
        elif c in (1, 2, 3, 7, 9):
            cell.alignment = center()
            cell.fill = fill(row_bg)
        else:
            cell.alignment = left()
            cell.fill = fill(row_bg)

    # Highlight PROVA/RETESTE row differently
    if tipo == 'PROVA':
        for c in range(1, 14):
            cell = ws1.cell(r, c)
            cell.fill = fill('FF5252')
            cell.font = Font(bold=True, size=11, color='FFFFFF')
    elif tipo == 'RETESTE':
        for c in range(1, 14):
            cell = ws1.cell(r, c)
            cell.fill = fill('2196F3')
            cell.font = Font(bold=True, size=11, color='FFFFFF')

    ws1.row_dimensions[r].height = 18

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 2 — PROGRESSÃO SEMANAL (para gráficos)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Progressão Semanal')
ws2.freeze_panes = 'A3'

ws2.merge_cells('A1:H1')
t2 = ws2['A1']
t2.value = 'PROGRESSÃO SEMANAL — USE ESTA ABA PARA GRÁFICOS'
t2.font = Font(bold=True, size=13, color='FFFFFF')
t2.fill = fill('1A3A6B')
t2.alignment = center()
ws2.row_dimensions[1].height = 26

headers2 = ['Semana','Início','Fase','Vol. Total (km)','Long Run (km)','Qualidade (km)','Leve/Moderado (km)','Semana Rec.']
col_widths2 = [9, 13, 17, 16, 14, 16, 20, 13]
for c, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(2, c, h)
    cell.fill = fill('2B5BA8')
    cell.font = hdr_font()
    cell.alignment = center()
    cell.border = thin_border()
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.row_dimensions[2].height = 22

# Aggregate by week
weeks = defaultdict(lambda: {'fase':'','data_inicio':'','sessions':[],'recovery':False})
for s in sessions:
    date_str, dia, sem, fase, tipo, dist, pace, zona, det, rec, *extra = s
    if sem == 0:
        continue  # skip warmup
    weeks[sem]['fase'] = fase
    weeks[sem]['sessions'].append((tipo, dist))
    weeks[sem]['recovery'] = rec
    if not weeks[sem]['data_inicio'] or date_str < weeks[sem]['data_inicio']:
        weeks[sem]['data_inicio'] = date_str

QUALITY_TYPES = {'TM','TC','TR','FK','PROG','RP','RP+TM','PROVA'}
EASY_TYPES    = {'CL','CM','CL+S','REC'}

for r, sem in enumerate(sorted(weeks.keys()), 3):
    w = weeks[sem]
    total = sum(d for _, d in w['sessions'])
    lr    = sum(d for t, d in w['sessions'] if t == 'LR')
    qual  = sum(d for t, d in w['sessions'] if t in QUALITY_TYPES)
    easy  = sum(d for t, d in w['sessions'] if t in EASY_TYPES)

    row_bg = 'FFF8E1' if w['recovery'] else FASE_BG.get(w['fase'], 'FFFFFF')
    vals = [f'S{sem:02d}', w['data_inicio'], w['fase'], round(total,1), round(lr,1), round(qual,1), round(easy,1), 'Sim' if w['recovery'] else '']

    for c, val in enumerate(vals, 1):
        cell = ws2.cell(r, c, val)
        cell.fill = fill(row_bg)
        cell.border = thin_border()
        cell.font = cell_font()
        cell.alignment = center() if c != 3 else left()
    ws2.row_dimensions[r].height = 18

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 3 — ZONAS DE TREINAMENTO
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Zonas de Treinamento')

ws3.merge_cells('A1:E1')
t3 = ws3['A1']
t3.value = 'ZONAS DE TREINAMENTO — VDOT 43 (Teste 5km: 26:00 · 03/mai/2026)'
t3.font = Font(bold=True, size=12, color='FFFFFF')
t3.fill = fill('1A3A6B')
t3.alignment = center()
ws3.row_dimensions[1].height = 24

headers3 = ['Zona','Nome','Pace Alvo','Percepção de Esforço','Uso Principal']
col_widths3 = [8, 22, 16, 35, 30]
for c, (h, w) in enumerate(zip(headers3, col_widths3), 1):
    cell = ws3.cell(2, c, h)
    cell.fill = fill('2B5BA8')
    cell.font = hdr_font()
    cell.alignment = center()
    cell.border = thin_border()
    ws3.column_dimensions[get_column_letter(c)].width = w
ws3.row_dimensions[2].height = 20

zones = [
    ('Z1','Recuperação',   '6:20–6:50/km','Totalmente confortável — conversa fácil', 'CL, LR, CL+S (parte leve)', 'BBDEFB'),
    ('Z2','Aeróbio Base',  '5:55–6:20/km','Conversa com esforço — frases longas',    'CM, parte do LR',            'C8E6C9'),
    ('Z3','Limiar / RP',   '5:25–5:35/km','Frases curtas — esforço sustentável',     'TR, RP, parte do PROG',      'FFE0B2'),
    ('Z4','VO2max',        '5:00–5:08/km','Difícil falar — intensidade alta',         'TM, TC',                     'FFCCBC'),
    ('Z5','Velocidade',    '4:40–4:52/km','Sem fala — esforço máximo',               'Strides (20s)',               'FFAB91'),
]
for r, (z, nome, pace, perc, uso, color) in enumerate(zones, 3):
    vals = [z, nome, pace, perc, uso]
    for c, val in enumerate(vals, 1):
        cell = ws3.cell(r, c, val)
        cell.fill = fill(color)
        cell.border = thin_border()
        cell.font = cell_font(bold=(c == 1))
        cell.alignment = center() if c in (1, 3) else left()
    ws3.row_dimensions[r].height = 20

# ═══════════════════════════════════════════════════════════════════════════════
# ABA 4 — LEGENDA DOS TREINOS
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Tipos de Treino')

ws4.merge_cells('A1:D1')
t4 = ws4['A1']
t4.value = 'TIPOS DE TREINO — LEGENDA'
t4.font = Font(bold=True, size=12, color='FFFFFF')
t4.fill = fill('1A3A6B')
t4.alignment = center()
ws4.row_dimensions[1].height = 24

headers4 = ['Sigla','Nome','Zona','Descrição']
col_widths4 = [10, 22, 10, 58]
for c, (h, w) in enumerate(zip(headers4, col_widths4), 1):
    cell = ws4.cell(2, c, h)
    cell.fill = fill('2B5BA8')
    cell.font = hdr_font()
    cell.alignment = center()
    cell.border = thin_border()
    ws4.column_dimensions[get_column_letter(c)].width = w
ws4.row_dimensions[2].height = 20

tipos = [
    ('LR',    'Corrida Longa',      'Z1-Z2', 'Maior volume semanal — base de resistência aeróbia',          'BBDEFB'),
    ('CL',    'Corrida Leve',       'Z1',    'Ritmo fácil e conversacional — recuperação ativa',             'C8E6C9'),
    ('CM',    'Corrida Moderada',   'Z2',    'Esforço controlado — consolida o volume semanal',              'DCEDC8'),
    ('CL+S',  'Leve + Strides',     'Z1+Z5', 'Corrida leve com 4–6 acelerações de 20s ao final',            'B2DFDB'),
    ('FK',    'Fartlek',            'Z2-Z4', 'Variações de ritmo informais — jogo de velocidade livre',      'E1BEE7'),
    ('PROG',  'Progressivo',        'Z2→Z3', 'Começa em Z2, termina em Z3 — simula a fadiga da prova',       'B2EBF2'),
    ('TM',    'Tiros Médios',       'Z4',    'Repetições de 1–1,5km — melhora VO2max e limiar',              'FFCCBC'),
    ('TC',    'Tiros Curtos',       'Z5',    'Repetições de 400–600m — velocidade e economia de corrida',    'FFAB91'),
    ('TR',    'Tempo Run',          'Z3',    'Ritmo contínuo sustentado por 20–45 min — limiar aeróbio',     'FFE0B2'),
    ('RP',    'Ritmo de Prova',     'Z3',    'Blocos no pace alvo da meia maratona (5:35–5:40/km)',           'FFCDD2'),
    ('RP+TM', 'RP + Tiros',         'Z3+Z4', 'Sessão mista: blocos de prova + tiros curtos de qualidade',    'FFCDD2'),
    ('PROVA', 'Meia Maratona',      'Z3',    '21,1km — Objetivo Sub-2h00 (pace 5:40/km)',                    'FF8A80'),
    ('REC',   'Recuperação',        'Z1',    'Trote muito leve pós-prova — regeneração ativa',               'EEEEEE'),
]
for r, (sig, nome, zona, desc, color) in enumerate(tipos, 3):
    vals = [sig, nome, zona, desc]
    for c, val in enumerate(vals, 1):
        cell = ws4.cell(r, c, val)
        cell.fill = fill(color)
        cell.border = thin_border()
        cell.font = cell_font(bold=(c == 1))
        cell.alignment = center() if c in (1, 3) else left()
    ws4.row_dimensions[r].height = 20

# ═══════════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT)
print(f'Arquivo salvo: {OUTPUT}')
print(f'Total de sessões: {len(sessions)}')
print(f'Total de semanas: {len(weeks)}')
